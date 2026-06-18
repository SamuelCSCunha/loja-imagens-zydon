#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Planilha entra -> planilha sai com a coluna de imagem preenchida com os links.

- Le TODAS as abas do .xlsx.
- Em cada aba, acha por CABECALHO a coluna com o nome do produto
  (produto/nome/item/...) e a coluna de imagem (imagem/foto/url/...).
  Se nao existir coluna de imagem, cria uma chamada "imagem".
- Para cada produto sem imagem, busca a 1a foto boa (sem chave), baixa,
  publica no repo (jsDelivr) e grava o link na celula.
- Celulas de imagem JA preenchidas sao puladas (pode rodar de novo p/ completar).
- Salva uma copia: <arquivo>_COM_IMAGENS.xlsx (nao mexe no original).

Uso:
    python planilha.py minha_planilha.xlsx
"""
import sys, os, re, time, subprocess, unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed

from ddgs import DDGS
from openpyxl import load_workbook

from buscar import slug, baixar_validar, BUSCA_MAX

IMG_DIR = "imagens"
WORKERS = 6

COLS_NOME = {"produto", "produtos", "nome", "nome do produto", "item",
             "descricao", "descricao do produto", "product", "title", "titulo"}
COLS_IMG = {"imagem", "imagens", "foto", "fotos", "url", "link", "imagem url",
            "image", "img", "url da imagem"}


def norm(v):
    if v is None:
        return ""
    t = unicodedata.normalize("NFKD", str(v)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", t).strip().lower()


def user_repo():
    r = subprocess.run(["git", "remote", "get-url", "origin"],
                       capture_output=True, text=True)
    if r.returncode != 0:
        sys.exit("ERRO: repo git sem 'origin'. Rode no diretorio da ferramenta.")
    m = re.search(r"github\.com[:/]([^/]+)/(.+?)(?:\.git)?$", r.stdout.strip())
    if not m:
        sys.exit(f"ERRO: remote nao reconhecido: {r.stdout.strip()}")
    return m.group(1), m.group(2)


def achar_imagem(produto):
    """Busca, valida e salva a 1a imagem boa em imagens/<slug>.jpg. Retorna slug ou None."""
    s = slug(produto)
    destino = os.path.join(IMG_DIR, f"{s}.jpg")
    if os.path.exists(destino):
        return s
    resultados = []
    for tentativa in range(3):                       # retry: DDG sem chave da rate-limit
        try:
            resultados = DDGS().images(produto, max_results=BUSCA_MAX)
            if resultados:
                break
        except Exception:
            pass
        time.sleep(1.5 * (tentativa + 1))
    if not resultados:
        print(f"  [sem resultado] {produto}")
        return None
    for item in resultados:
        url = item.get("image")
        if not url:
            continue
        out = baixar_validar(url)
        if out:
            with open(destino, "wb") as f:
                f.write(out[0])
            return s
    return None


def coletar_tarefas(wb):
    """Varre abas, detecta colunas, retorna (tarefas, plano_por_aba)."""
    tarefas = {}   # produto -> None (slug depois)
    plano = []     # (ws, nome_col, img_col, [(row, produto)])
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        nome_col = img_col = None
        for col in range(1, ws.max_column + 1):
            h = norm(ws.cell(1, col).value)
            if not nome_col and h in COLS_NOME:
                nome_col = col
            elif not img_col and h in COLS_IMG:
                img_col = col
        if not nome_col:
            print(f"  [pula aba '{ws.title}'] sem coluna de nome (produto/nome/...)")
            continue
        if not img_col:
            img_col = ws.max_column + 1
            ws.cell(1, img_col, "imagem")
        itens = []
        for row in range(2, ws.max_row + 1):
            produto = ws.cell(row, nome_col).value
            if produto is None or not str(produto).strip():
                continue
            if ws.cell(row, img_col).value:   # ja tem link -> pula
                continue
            produto = str(produto).strip()
            itens.append((row, produto))
            tarefas[produto] = None
        if itens:
            plano.append((ws, nome_col, img_col, itens))
            print(f"  aba '{ws.title}': {len(itens)} produto(s) sem imagem")
    return tarefas, plano


def main():
    if len(sys.argv) < 2 or not os.path.exists(sys.argv[1]):
        sys.exit("Uso: python planilha.py <arquivo.xlsx>")
    caminho = sys.argv[1]
    if not caminho.lower().endswith(".xlsx"):
        sys.exit("So aceito .xlsx (com abas). Exporte sua planilha para .xlsx.")

    os.makedirs(IMG_DIR, exist_ok=True)
    wb = load_workbook(caminho)
    tarefas, plano = coletar_tarefas(wb)
    if not tarefas:
        print("Nada a fazer: todas as celulas de imagem ja preenchidas (ou sem produtos).")
        return

    print(f"\nBuscando {len(tarefas)} produto(s) unico(s) com {WORKERS} simultaneas...")
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = {ex.submit(achar_imagem, p): p for p in tarefas}
        for fut in as_completed(futs):
            p = futs[fut]
            tarefas[p] = fut.result()
    achados = {p: s for p, s in tarefas.items() if s}
    faltou = [p for p, s in tarefas.items() if not s]
    print(f"Encontrei imagem para {len(achados)}/{len(tarefas)}.")

    if achados:
        subprocess.run(["git", "add", IMG_DIR], capture_output=True, text=True)
        c = subprocess.run(["git", "commit", "-m",
                            f"imagens: +{len(achados)} produto(s) via planilha"],
                           capture_output=True, text=True)
        p = subprocess.run(["git", "push", "origin", "HEAD"],
                           capture_output=True, text=True)
        if p.returncode != 0:
            sys.exit(f"ERRO no push (links nao funcionariam): {p.stderr}")

    user, repo = user_repo()
    base = f"https://cdn.jsdelivr.net/gh/{user}/{repo}@main/{IMG_DIR}"
    escritos = 0
    for ws, nome_col, img_col, itens in plano:
        for row, produto in itens:
            s = tarefas.get(produto)
            if s:
                ws.cell(row, img_col, f"{base}/{s}.jpg")
                escritos += 1

    stem = re.sub(r"_COM_IMAGENS$", "", os.path.splitext(caminho)[0])
    saida = stem + "_COM_IMAGENS.xlsx"
    wb.save(saida)
    print(f"\nGravei {escritos} link(s). Planilha: {saida}")
    if faltou:
        print(f"SEM imagem (revise/refine o nome): {', '.join(faltou)}")


if __name__ == "__main__":
    main()
