#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
App local (navegador) para popular planilhas com imagens de produto.
Tudo por tela: arrasta o .xlsx -> ve o progresso -> revisa/troca -> baixa.

Nao chame direto: use o "Gerador de Imagens.bat" (duplo-clique).
"""
import os, re, io, csv, time, uuid, shutil, threading, subprocess, webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed

from flask import (Flask, request, jsonify, send_file, send_from_directory,
                   render_template_string, abort)
from ddgs import DDGS
from openpyxl import load_workbook

from buscar import slug, baixar_validar, BUSCA_MAX

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WORK_DIR = os.path.join(BASE_DIR, "work")
IMG_DIR = os.path.join(BASE_DIR, "imagens")
N_CANDIDATOS = 3
WORKERS = 6

COLS_NOME = {"produto", "produtos", "nome", "nome do produto", "item",
             "descricao", "descricao do produto", "product", "title", "titulo"}
COLS_IMG = {"imagem", "imagens", "foto", "fotos", "url", "link", "imagem url",
            "image", "img", "url da imagem"}

app = Flask(__name__)
JOBS = {}
import unicodedata


def norm(v):
    if v is None:
        return ""
    t = unicodedata.normalize("NFKD", str(v)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", t).strip().lower()


def user_repo():
    r = subprocess.run(["git", "remote", "get-url", "origin"],
                       capture_output=True, text=True, cwd=BASE_DIR)
    if r.returncode != 0:
        return None, None
    m = re.search(r"github\.com[:/]([^/]+)/(.+?)(?:\.git)?$", r.stdout.strip())
    return (m.group(1), m.group(2)) if m else (None, None)


def buscar_top3(produto, outdir, n=N_CANDIDATOS):
    s = slug(produto)
    resultados = []
    for t in range(3):
        try:
            resultados = DDGS().images(produto, max_results=BUSCA_MAX)
            if resultados:
                break
        except Exception:
            pass
        time.sleep(1.5 * (t + 1))
    urls = [it["image"] for it in resultados if it.get("image")][:10]
    with ThreadPoolExecutor(max_workers=8) as ex:   # baixa candidatos em paralelo
        outs = list(ex.map(baixar_validar, urls))
    cands = []
    for out in outs:
        if len(cands) >= n:
            break
        if not out:
            continue
        data, w, h = out
        nome = f"{s}__{len(cands)+1}.jpg"
        with open(os.path.join(outdir, nome), "wb") as f:
            f.write(data)
        cands.append({"arquivo": nome, "w": w, "h": h})
    return {"produto": produto, "slug": s, "candidatos": cands}


def coletar(wb):
    prodslug, plano, unicos = {}, [], []
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        nome_col = img_col = None
        for col in range(1, ws.max_column + 1):
            h = norm(ws.cell(1, col).value)
            if not nome_col and h in COLS_NOME:
                nome_col = col
            elif not img_col and h in COLS_IMG:
                img_col = col
        if not nome_col:
            continue
        if not img_col:
            img_col = ws.max_column + 1
            ws.cell(1, img_col, "imagem")
        itens = []
        for row in range(2, ws.max_row + 1):
            produto = ws.cell(row, nome_col).value
            if produto is None or not str(produto).strip():
                continue
            if ws.cell(row, img_col).value:
                continue
            produto = str(produto).strip()
            itens.append((row, produto))
            if produto not in prodslug:
                prodslug[produto] = slug(produto)
                unicos.append(produto)
        if itens:
            plano.append((ws, nome_col, img_col, itens))
    return prodslug, plano, unicos


def worker(job_id):
    job = JOBS[job_id]
    try:
        wb = load_workbook(job["input"])
        prodslug, plano, unicos = coletar(wb)
        job.update(wb=wb, plano=plano, prodslug=prodslug,
                   total=len(unicos), done=0, products=[])
        cdir = os.path.join(job["dir"], "candidatos")
        os.makedirs(cdir, exist_ok=True)
        if not unicos:
            job["status"] = "review"
            return
        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            futs = {ex.submit(buscar_top3, p, cdir): p for p in unicos}
            for fut in as_completed(futs):
                job["products"].append(fut.result())
                job["done"] += 1
        job["status"] = "review"
    except Exception as e:
        job["status"] = "erro"
        job["erro"] = str(e)


PAGE = r"""<!doctype html><html lang=pt-br><head><meta charset=utf-8>
<title>Gerador de Imagens de Produto</title><style>
*{box-sizing:border-box} body{font-family:system-ui,Arial;margin:0;background:#0f1115;color:#e8e8e8}
header{background:#161a22;padding:16px 24px;border-bottom:1px solid #232a36}
h1{font-size:18px;margin:0} .wrap{max-width:1100px;margin:0 auto;padding:24px}
#drop{border:2px dashed #3a4252;border-radius:14px;padding:48px;text-align:center;
background:#141821;cursor:pointer;transition:.15s}
#drop.hi{border-color:#5b8cff;background:#1b2236} .muted{color:#8b95a7}
button{background:#5b8cff;border:0;color:#fff;padding:11px 20px;border-radius:9px;
font-weight:600;cursor:pointer;font-size:15px} button.sec{background:#2a3140}
button:disabled{opacity:.5;cursor:default}
.bar{height:10px;background:#222a38;border-radius:6px;overflow:hidden;margin:18px 0}
.bar>i{display:block;height:100%;background:#5b8cff;width:0;transition:.3s}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(230px,1fr));gap:16px;margin-top:18px}
.card{background:#161a22;border:1px solid #232a36;border-radius:12px;padding:12px}
.card h3{font-size:13px;margin:0 0 8px;line-height:1.3;min-height:34px}
.opts{display:flex;gap:6px;flex-wrap:wrap}
.opt{border:2px solid #2a2f3a;border-radius:8px;padding:3px;cursor:pointer;background:#fff}
.opt img{width:60px;height:60px;object-fit:contain;display:block}
.opt.sel{border-color:#5b8cff} .none{color:#8b95a7;font-size:12px;width:60px;height:60px;
display:flex;align-items:center;justify-content:center;border:2px solid #2a2f3a;border-radius:8px;cursor:pointer}
.none.sel{border-color:#ff6b6b;color:#ff6b6b}
.redo{margin-top:8px;display:flex;gap:6px} .redo input{flex:1;background:#0f1115;border:1px solid #2a2f3a;
color:#e8e8e8;border-radius:7px;padding:6px 8px;font-size:12px}
.sem{border-color:#5a2a2a} .toolbar{display:flex;gap:12px;align-items:center;margin:20px 0;flex-wrap:wrap}
a.dl{display:inline-block;background:#2fbf71;color:#fff;padding:11px 20px;border-radius:9px;
font-weight:600;text-decoration:none} #msg{margin-left:auto;color:#8b95a7}
small{color:#8b95a7}
</style></head><body>
<header><h1>🖼️ Gerador de Imagens de Produto</h1>
<small class=muted>planilha entra → coluna de imagem sai com os links</small></header>
<div class=wrap>
<div id=step1>
  <div id=drop><b>Arraste sua planilha .xlsx aqui</b><br><span class=muted>ou clique para escolher · várias abas, ok</span>
  <input id=file type=file accept=".xlsx" hidden></div>
</div>
<div id=step2 style=display:none>
  <div class=muted id=plabel>Buscando imagens…</div>
  <div class=bar><i id=pbar></i></div>
</div>
<div id=step3 style=display:none>
  <div class=toolbar>
    <button id=confirm>Gerar links e baixar planilha</button>
    <span id=msg></span>
  </div>
  <div class=grid id=grid></div>
</div>
</div>
<script>
const $=s=>document.querySelector(s);
let JOB=null, PROD={}, SEL={};
const drop=$('#drop'), file=$('#file');
drop.onclick=()=>file.click();
drop.ondragover=e=>{e.preventDefault();drop.classList.add('hi')};
drop.ondragleave=()=>drop.classList.remove('hi');
drop.ondrop=e=>{e.preventDefault();drop.classList.remove('hi');if(e.dataTransfer.files[0])upload(e.dataTransfer.files[0])};
file.onchange=()=>{if(file.files[0])upload(file.files[0])};

async function upload(f){
  if(!f.name.toLowerCase().endsWith('.xlsx')){alert('Só .xlsx (exporte do Google Sheets como Excel).');return;}
  const fd=new FormData();fd.append('file',f);
  $('#step1').style.display='none';$('#step2').style.display='';
  const r=await fetch('/upload',{method:'POST',body:fd});const j=await r.json();
  if(j.erro){alert(j.erro);location.reload();return;}
  JOB=j.job;poll();
}
async function poll(){
  const r=await fetch('/status?job='+JOB);const j=await r.json();
  if(j.status==='erro'){alert('Erro: '+j.erro);location.reload();return;}
  const pct=j.total?Math.round(100*j.done/j.total):100;
  $('#pbar').style.width=pct+'%';
  $('#plabel').textContent=`Buscando imagens…  ${j.done}/${j.total}`;
  if(j.status==='review'){render(j.products);return;}
  setTimeout(poll,800);
}
function render(products){
  $('#step2').style.display='none';$('#step3').style.display='';
  if(!products.length){$('#grid').innerHTML='<p class=muted>Nada a preencher: todas as células de imagem já tinham link (ou não achei coluna de produto).</p>';$('#confirm').disabled=true;return;}
  const g=$('#grid');g.innerHTML='';
  for(const p of products){PROD[p.slug]=p;card(g,p);}
}
function card(g,p){
  const d=document.createElement('div');d.className='card'+(p.candidatos.length?'':' sem');d.id='c_'+p.slug;
  let opts='';
  p.candidatos.forEach((c,i)=>{opts+=`<div class=opt data-f="${c.arquivo}" onclick="pick('${p.slug}','${c.arquivo}',this)"><img src="/cand/${JOB}/${c.arquivo}"></div>`;});
  opts+=`<div class=none onclick="pick('${p.slug}','',this)">sem</div>`;
  d.innerHTML=`<h3>${p.produto}</h3><div class=opts>${opts}</div>
   <div class=redo><input placeholder="buscar com outro nome…" id="q_${p.slug}">
   <button class=sec onclick="redo('${p.slug}')">↻</button></div>`;
  g.appendChild(d);
  if(p.candidatos.length){pick(p.slug,p.candidatos[0].arquivo,d.querySelector('.opt'));}
  else{SEL[p.slug]='';d.querySelector('.none').classList.add('sel');}
}
function pick(slug,arq,el){
  SEL[slug]=arq;
  const c=$('#c_'+slug);c.querySelectorAll('.opt,.none').forEach(o=>o.classList.remove('sel'));
  el.classList.add('sel');
}
async function redo(slug){
  const q=$('#q_'+slug).value.trim();if(!q)return;
  const c=$('#c_'+slug);c.style.opacity=.5;
  const r=await fetch('/research',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({job:JOB,slug,query:q})});
  const p=await r.json();c.style.opacity=1;
  PROD[slug]=p;const g=document.createElement('div');card(g,p);c.replaceWith(g.firstChild);
}
$('#confirm').onclick=async()=>{
  $('#confirm').disabled=true;$('#msg').textContent='publicando imagens e gerando links…';
  const r=await fetch('/confirm',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({job:JOB,sel:SEL})});
  const j=await r.json();
  if(j.erro){$('#msg').textContent='Erro: '+j.erro;$('#confirm').disabled=false;return;}
  $('#msg').innerHTML=`✅ ${j.gravados} link(s) gravados. `+
    `<a class=dl href="/download?job=${JOB}">⬇ Baixar planilha</a>`+
    (j.faltou&&j.faltou.length?`<br><small>sem imagem: ${j.faltou.join(', ')}</small>`:'');
};
</script></body></html>"""


@app.route("/")
def index():
    return render_template_string(PAGE)


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f or not f.filename.lower().endswith(".xlsx"):
        return jsonify(erro="Envie um arquivo .xlsx")
    job_id = uuid.uuid4().hex[:10]
    jdir = os.path.join(WORK_DIR, job_id)
    os.makedirs(jdir, exist_ok=True)
    inp = os.path.join(jdir, "input.xlsx")
    f.save(inp)
    JOBS[job_id] = {"dir": jdir, "input": inp, "status": "buscando",
                    "done": 0, "total": 0, "products": []}
    threading.Thread(target=worker, args=(job_id,), daemon=True).start()
    return jsonify(job=job_id)


@app.route("/status")
def status():
    job = JOBS.get(request.args.get("job"))
    if not job:
        return jsonify(status="erro", erro="job inexistente")
    return jsonify(status=job["status"], done=job["done"], total=job["total"],
                   products=job["products"] if job["status"] == "review" else [],
                   erro=job.get("erro"))


@app.route("/cand/<job>/<path:fn>")
def cand(job, fn):
    j = JOBS.get(job)
    if not j:
        abort(404)
    return send_from_directory(os.path.join(j["dir"], "candidatos"), fn)


@app.route("/research", methods=["POST"])
def research():
    d = request.get_json()
    job = JOBS.get(d.get("job"))
    if not job:
        return jsonify(erro="job inexistente"), 400
    cdir = os.path.join(job["dir"], "candidatos")
    p = buscar_top3(d["query"], cdir)
    p["slug"] = d["slug"]                      # mantem o slug original do produto
    # renomeia arquivos para o slug original
    fixed = []
    for i, c in enumerate(p["candidatos"], 1):
        novo = f"{d['slug']}__r{i}.jpg"
        os.replace(os.path.join(cdir, c["arquivo"]), os.path.join(cdir, novo))
        c["arquivo"] = novo
        fixed.append(c)
    p["candidatos"] = fixed
    p["produto"] = next((x["produto"] for x in job["products"]
                         if x["slug"] == d["slug"]), d["query"])
    for x in job["products"]:
        if x["slug"] == d["slug"]:
            x["candidatos"] = fixed
    return jsonify(p)


@app.route("/confirm", methods=["POST"])
def confirm():
    d = request.get_json()
    job = JOBS.get(d.get("job"))
    if not job:
        return jsonify(erro="job inexistente"), 400
    sel = d.get("sel", {})
    cdir = os.path.join(job["dir"], "candidatos")
    os.makedirs(IMG_DIR, exist_ok=True)
    publicados = {}
    for s, arq in sel.items():
        if not arq:
            continue
        src = os.path.join(cdir, arq)
        if os.path.exists(src):
            shutil.copyfile(src, os.path.join(IMG_DIR, f"{s}.jpg"))
            publicados[s] = True

    if publicados:
        subprocess.run(["git", "add", "imagens"], cwd=BASE_DIR,
                       capture_output=True, text=True)
        subprocess.run(["git", "commit", "-m",
                        f"imagens: +{len(publicados)} via app"], cwd=BASE_DIR,
                       capture_output=True, text=True)
        push = subprocess.run(["git", "push", "origin", "HEAD"], cwd=BASE_DIR,
                              capture_output=True, text=True)
        if push.returncode != 0:
            return jsonify(erro="falha ao publicar no GitHub: " + push.stderr[-200:])

    user, repo = user_repo()
    if not user:
        return jsonify(erro="repositorio GitHub nao configurado")
    base = f"https://cdn.jsdelivr.net/gh/{user}/{repo}@main/imagens"

    wb = job["wb"]
    prodslug = job["prodslug"]
    gravados, faltou = 0, []
    for ws, nome_col, img_col, itens in job["plano"]:
        for row, produto in itens:
            s = prodslug.get(produto)
            if s in publicados:
                ws.cell(row, img_col, f"{base}/{s}.jpg")
                gravados += 1
    for s, arq in sel.items():
        if not arq:
            prod = next((x["produto"] for x in job["products"] if x["slug"] == s), s)
            faltou.append(prod)

    out = os.path.join(job["dir"], "saida.xlsx")
    wb.save(out)
    job["output"] = out
    return jsonify(gravados=gravados, faltou=faltou)


@app.route("/download")
def download():
    job = JOBS.get(request.args.get("job"))
    if not job or not job.get("output"):
        abort(404)
    return send_file(job["output"], as_attachment=True,
                     download_name="planilha_COM_IMAGENS.xlsx")


def abrir_navegador():
    time.sleep(1.2)
    webbrowser.open("http://127.0.0.1:5000")


if __name__ == "__main__":
    os.makedirs(WORK_DIR, exist_ok=True)
    threading.Thread(target=abrir_navegador, daemon=True).start()
    print("\n  Gerador de Imagens rodando em http://127.0.0.1:5000")
    print("  (deixe esta janela aberta enquanto usa; feche para encerrar)\n")
    app.run(host="127.0.0.1", port=5000, threaded=True)
